import subprocess
from openpyxl import Workbook
import openpyxl
try:
    profilleri_al = subprocess.run(['netsh', 'wlan', 'show', 'profiles'], stdout = subprocess.PIPE)
    profillerhepsi = profilleri_al.stdout.decode('latin1')
    profiller = []

    for satır in profillerhepsi.split('\n'):
        if 'User Profile' in satır:
            profilayir = satır.split(':')[1]
            profil = profilayir[1:]
            profiller.append(profil)
    wb = Workbook()
    ws = wb.active
    ws.title = "Deneme"
    ws = wb.create_sheet("Şifrelerrr")
    print(wb.sheetnames)

    for profil in profiller:
        profildetay = subprocess.run(['netsh', 'wlan', 'show', 'profile', profil[:-1], 'key=clear'], stdout = subprocess.PIPE)
        detaylar = profildetay.stdout.decode('latin1')
        for satır in detaylar.split('\n'):
            if 'Key Content' in satır:
                bosluklusifre = satır.split(':')[1]
                sifre = bosluklusifre[1:]             
                ws.append([profil , sifre])
    dosyayolu = ("C:/Users/cagla/Desktop/odevv/Sifrelerr.xlsx")            
    wb.save(dosyayolu)  

except:
    print("olmadı")
